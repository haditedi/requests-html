import time
import sys
import openpyxl
from datetime import datetime, timedelta
from requests_html import AsyncHTMLSession

#getting and validating input for start date and end date (should not exceed 1 year)
max_date = timedelta(days=366)
todayDate = datetime.now()
while True:
	print("Please enter the start date (dd-mm-yy): ")
	inStartDate = input()
	try:
		start_date = datetime.strptime(inStartDate[:2]+'-'+inStartDate[3:5]+'-'+inStartDate[6:8], '%d-%m-%y')
		if start_date > todayDate and start_date < (todayDate+max_date):
			break
		else:
			print("date should be at least tommorrow and not exceed 1 year")
	except Exception as e:
		print(e)
while True:
	print("Please enter end date (dd-mm-yy): ")
	inEndDate = input()
	try:
		end_Date_User = datetime.strptime(inEndDate[:2]+'-'+inEndDate[3:5]+'-'+inEndDate[6:8], '%d-%m-%y')
		if end_Date_User < start_date:
			print("End date should be greater than start date")
		elif end_Date_User > (start_date+max_date):
			print("End date should not be more than a year")
		else:
			break
	except Exception as e:
		print(e)
print("processing,,,\n")
#-----------------

#Storing formatted date string in a list. Date requests will be every 3 days and length of stay is 10 nights
#these are for Ashburn Court Apartments 
ascstart_date = start_date
ascdep_date = end_Date_User
ascstart_url=[]
ascdep_url=[]
num_nights = timedelta(days=10)
numdays = timedelta(days=3)
for i in range(365):
	if ascstart_date >= end_Date_User:
		ascstart_url.append(end_Date_User.strftime("%Y-%m-%d"))
		ascdep_date = end_Date_User + num_nights
		ascdep_url.append(ascdep_date.strftime("%Y-%m-%d"))
		break
	else:
		ascstart_url.append(ascstart_date.strftime("%Y-%m-%d"))
		ascdep_date = ascstart_date + num_nights
		ascdep_url.append(ascdep_date.strftime("%Y-%m-%d"))
		ascstart_date += numdays
#------------

#Storing formatted date string in a list. Date requests will be every 3 days and length of stay is 10 nights
#these are for Cheval Harrington Court 
chcstart_date =  start_date
chcstart_url=[]
for i in range(365):
	if chcstart_date >= end_Date_User:
		chcstart_url.append(end_Date_User.strftime("%Y-%m-%d"))
		break
	else:
		chcstart_url.append(chcstart_date.strftime("%Y-%m-%d"))
		chcstart_date += numdays
chcnum_nights = 10
#-------------


#Set up variables for the excel file
wb = openpyxl.load_workbook('multi.xlsx')
sheet = wb.active
sheet.cell(row=1, column=3).value = "DATE"
sheet.cell(row=1, column=4).value = "CHC-1 BED 548ft2"
sheet.cell(row=1, column=5).value = "CHC-2 BED 656ft2"
sheet.cell(row=1, column=6).value = "ASC-1 BED 624ft2"
sheet.cell(row=1, column=7).value = "ASC-2 BED 936ft2"
sheet.cell(row=1, column=8).value = "ASC-3 BED 624ft2"
thelast=sheet.max_row

#Requesting data from Cheval Harrington Court and Ashburn Court
for i in range(len(ascstart_url)):
	ascdate=ascstart_url[i]
	ascdep_date = ascdep_url[i]	

	asession = AsyncHTMLSession()

	async def getasc():
		r = await asession.get(f"https://booking.maykenbel.com/?chain=19159&template= \
			maykenbel&shell=MKNBL2018&start=availresults&brand=maykenbe&currency=GBP&lang= \
			1&arrive={ascdate[5:7]}%2F{ascdate[8:10]}%2F{ascdate[:4]}&depart= \
			{ascdep_date[5:7]}%2F{ascdep_date[8:10]}%2F{ascdep_date[:4]}&hotel= \
			70825&dpArrive={ascdate[8:10]}%2F{ascdate[5:7]}%2F{ascdate[:4]}&dpDepart= \
			{ascdep_date[8:10]}%2F{ascdep_date[5:7]}%2F{ascdep_date[:4]}&rooms=1&adult=1&promo=")
		return r

	async def getchc():
		r = await asession.get(f"https://secure.chevalresidences.com/portal/site/www.chevalresidences.com/en/results.php?checkin= \
			{chcstart_url[i]}&nights={chcnum_nights}&keyword=CHC")
		return r
		
	results = asession.run(getasc, getchc)
	
	print("Date " + ascstart_url[i])
	sheet.cell(row=thelast+1+i, column=3).value = ascstart_url[i]
	

	#Storing requested datas
	try:
		chc1bed = results[1].html.find("span[id*='mbprice_'][id$='15070']", first=True).text
	except:
		print("No availability -CHC- Superior One Bedroom")
	else:
		price1bed_exvat=chc1bed.replace(",","")
		nprice1bed=(float(price1bed_exvat))/1.2
		nprice1bed=round(nprice1bed/chcnum_nights)
		print("Superior One Bedroom -CHC- " + "£ "+ str(nprice1bed))
		sheet.cell(row=thelast+1+i, column=4).value = nprice1bed
		
	try:
		chc2bed = results[1].html.find("span[id*='mbprice_'][id$='15071']", first=True).text
	except:
		print("No availability -CHC- 2 Bedroom Apartment")
	else:
		price2bed_exvat=chc2bed.replace(",","")
		nprice2bed=(float(price2bed_exvat))/1.2
		nprice2bed=round(nprice2bed/chcnum_nights)	
		print("2 Bedroom Apartment -CHC- " + "£ "+ str(nprice2bed))
		sheet.cell(row=thelast+1+i, column=5).value = nprice2bed
		
	#arranging datas returned from Ashburn Court
	try:
		asc1bed = results[0].html.find("div.ProductsList div[data-room-code='A1F'] span[id*='PriceData']", first=True).text
	except:
		print("No availability -ASC- Deluxe 1 Bed")
	else:
		ascPrice1bed_exvat=asc1bed.replace(",","")
		ascPrice1bed=(float(ascPrice1bed_exvat))/1.2
		ascPrice1bed=round(ascPrice1bed)
		sheet.cell(row=thelast+1+i, column=6).value = ascPrice1bed
		print("Deluxe 1 bedroom -ASC- £ " + str(ascPrice1bed))
	try:
		asc2bed = results[0].html.find("div.ProductsList div[data-room-code='2BD'] span[id*='PriceData']", first=True).text
	except:
		print("No availability -ASC- Deluxe 2 Bed")
	else:
		ascPrice2bed_exvat=asc2bed.replace(",","")
		ascPrice2bed=(float(ascPrice2bed_exvat))/1.2
		ascPrice2bed=round(ascPrice2bed)
		sheet.cell(row=thelast+1+i, column=7).value = ascPrice2bed
		print("Deluxe 2 bedroom -ASC- £ " + str(ascPrice2bed))
	try:
		asc3bed = results[0].html.find("div.ProductsList div[data-room-code='3BD'] span[id*='PriceData']", first=True).text
	except:
		print("No availability -ASC- Deluxe 3 Bed")
	else:
		ascPrice3bed_exvat=asc3bed.replace(",","")
		ascPrice3bed=(float(ascPrice3bed_exvat))/1.2
		ascPrice3bed=round(ascPrice3bed)
		sheet.cell(row=thelast+1+i, column=8).value = ascPrice3bed
		print("Deluxe 3 bedroom- ASC- £ " + str(ascPrice3bed))


	print("")
	time.sleep(70)

wb.save('multi.xlsx')





