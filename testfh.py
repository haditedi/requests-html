import time
import sys
import openpyxl
from datetime import datetime, timedelta
from requests_html import HTMLSession



todayDate = datetime.now()

#getting and validating input for start date and end date (should not exceed 1 year)
while True:
	print("Please enter the start date (dd-mm-yy): ")
	inStartDate = input()
	try:
		start_date = datetime.strptime(inStartDate[:2]+'-'+inStartDate[3:5]+'-'+inStartDate[6:8], '%d-%m-%y')
		if start_date > todayDate:
			break
		else:
			print("date should be at least tommorrow")
	except Exception as e:
		print(e)
while True:
	max_date = timedelta(days=366)
	print("Please enter end date (dd-mm-yy): ")
	inEndDate = input()
	try:
		end_Date_User = datetime.strptime(inEndDate[:2]+'-'+inEndDate[3:5]+'-'+inEndDate[6:8], '%d-%m-%y')
		if end_Date_User < start_date:
			print("End date should be greater than start date")
		elif end_Date_User > (start_date+max_date):
			print("End date should not be more than a year")
		else:
			break
	except Exception as e:
		print(e)


print("processing,,,\n")


'''
#Storing formatted date string in a list. Date requests will be every 3 days and length of stay is 10 
start_url=[]
for i in range(365):
	days = timedelta(days=3)
	if start_date >= end_Date_User:
		start_url.append(end_Date_User.strftime("%Y-%m-%d"))
		break
	else:
		start_url.append(start_date.strftime("%Y-%m-%d"))
		start_date = start_date + days
'''

#Storing url in a list for Fountain House



fhstart_url=[]
for i in range(365):
	days = timedelta(days=3)
	stay = timedelta(days=20)
	if start_date >= end_Date_User:
		fhday = end_Date_User.day
		fhmonth = end_Date_User.month
		fhyear = end_Date_User.year
		fhaddress = f"https://booking.maykenbel.com/?chain=19159&template=maykenbel&shell=MKNBL2018&start= \
				    availresults&brand=maykenbe&currency=GBP&lang=1&arrive={fhmonth}%2F{fhday}%2F{fhyear}&depart={fhmonth}%2F{fhday}%2F{fhyear}&hotel=\
				    70831&dpArrive={fhday}%2F{fhmonth}%2F{fhyear}&dpDepart={fhday}%2F{fhmonth}%2F{fhyear}&rooms=1&adult=1&promo="
		fhstart_url.append(fhaddress)
		break
	else:
		fhday = start_date.day
		fhmonth = start_date.month
		fhyear = start_date.year
		fhDday = start_date.day + stay
		fhDmonth = start_date.month + stay
		fhDyear = start_date.year + stay
		fhaddress= f"https://booking.maykenbel.com/?chain=19159&template=maykenbel&shell=MKNBL2018&start= \
				   availresults&brand=maykenbe&currency=GBP&lang=1&arrive={fhmonth}%2F{fhday}%2F{fhyear}&depart=\
				   {fhDmonth}%2F{fhDday}%2F{fhDyear}&hotel=\
				   70831&dpArrive={fhday}%2F{fhmonth}%2F{fhyear}&dpDepart={fhDday}%2F{fhDmonth}%2F{fhDyear}&rooms=1&adult=1&promo="
		fhstart_url.append(fhaddress)
		start_date = start_date + days

for i in range(len(fhstart_url)):
	print(str(i) + fhstart_url[i])
	print("")

#Time to go out and get those datas and save to an excel file
'''
#setup excel
wb = openpyxl.load_workbook('cheval.xlsx')
sheet = wb.active
sheet.cell(row=1, column=3).value = "DATE"
sheet.cell(row=1, column=4).value = "CHC-SUPERIOR 1 BED"
sheet.cell(row=1, column=5).value = "CHC-TWO BEDROOM"
thelast=sheet.max_row

num_nights = 10


for i in range(len(start_url)):
	try:
		session = HTMLSession()
		r = session.get("https://secure.chevalresidences.com/portal/site/www.chevalresidences.com/en/results.php?checkin={start}&nights={num}&keyword=CHC".format(start=start_url[i], num=num_nights))
		r.html.render()

		print("Date " + start_url[i])
		sheet.cell(row=thelast+1+i, column=3).value = start_url[i]

		try:
			price1bed = r.html.find("span[id*='mbprice_'][id$='15070']", first=True).text
		except:
			print("No availability - Superior One Bedroom")
		else:
			price1bed_exvat=price1bed.replace(",","")
			nprice1bed=(float(price1bed_exvat))/1.2
			nprice1bed=round(nprice1bed/num_nights)
			print("Superior One Bedroom " + "£ "+ str(nprice1bed))
			sheet.cell(row=thelast+1+i, column=4).value = nprice1bed

		try:
			price2bed = r.html.find("span[id*='mbprice_'][id$='15071']", first=True).text
		except:
			print("No availability - 2 Bedroom Apartment")
		else:
			price2bed_exvat=price2bed.replace(",","")
			nprice2bed=(float(price2bed_exvat))/1.2
			nprice2bed=round(nprice2bed/num_nights)	
			print("2 Bedroom Apartment " + "£ "+ str(nprice2bed))
			sheet.cell(row=thelast+1+i, column=5).value = nprice2bed
			
		
		print("")
		time.sleep(70)

	except KeyboardInterrupt:
		print("exit")
		sys.exit()
	except Exception as e:
		print(e)

wb.save('cheval.xlsx')
'''