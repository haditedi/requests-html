import time
import sys
import openpyxl
from datetime import datetime, timedelta
from requests_html import HTMLSession

# A program to fetch rate from a serviced apartments provider and is applicable for hotel as well

# Getting and validating input for start date and end date (should not exceed 1 year)
maxDate = timedelta(days=366)
todayDate = datetime.now()
while True:
	print("Please enter the start date (dd-mm-yy): ")
	inStartDate = input()
	try:
		start_date = datetime.strptime(inStartDate[:2]+'-'+inStartDate[3:5]+'-'+inStartDate[6:8], '%d-%m-%y')
		if start_date > todayDate and start_date < (todayDate+maxDate):
			break
		else:
			print("date should be at least tommorrow and not exceed 1 year")
	except Exception as e:
		print(e)
while True:
	print("Please enter end date (dd-mm-yy): ")
	inEndDate = input()
	try:
		end_Date_User = datetime.strptime(inEndDate[:2]+'-'+inEndDate[3:5]+'-'+inEndDate[6:8], '%d-%m-%y')
		if end_Date_User < start_date:
			print("End date should be greater than start date")
		elif end_Date_User > (start_date+maxDate):
			print("End date should not be more than a year")
		else:
			break
	except Exception as e:
		print(e)
print("processing,,,\n")
# -----------------------


# Storing formatted date string in a list. Date requests will be every 3 days and length of stay is 10 nights 
start_url=[]
for i in range(365):
	days = timedelta(days=3)
	if start_date >= end_Date_User:
		start_url.append(end_Date_User.strftime("%Y-%m-%d"))
		break
	else:
		start_url.append(start_date.strftime("%Y-%m-%d"))
		start_date = start_date + days
num_nights = 10 

# Setting up excel sheets
destFilename = 'rateCompare.xlsx'
try:
	wb = openpyxl.load_workbook(filename=destFilename)
except Exception as e:
	wb = openpyxl.Workbook()

sheet = wb.active
sheet.cell(row=1, column=3).value = "DATE"
sheet.cell(row=1, column=4).value = "CHC-SUPERIOR 1 BED"
sheet.cell(row=1, column=5).value = "CHC-TWO BEDROOM"
thelast=sheet.max_row

# Going out to fetch the data as inputted earlier and pasting the data in Excel file
for i in range(len(start_url)):
	try:
		session = HTMLSession()
		r = session.get("https://secure.chevalresidences.com/portal/site/www.chevalresidences.com/en/results.php?checkin={start}&nights={num}&keyword=CHC".format(start=start_url[i], num=num_nights))
		r.html.render()

		print("Date " + start_url[i])
		sheet.cell(row=thelast+1+i, column=3).value = start_url[i]

		try:
			price1bed = r.html.find("span[id*='mbprice_'][id$='15070']", first=True).text
		except:
			print("No availability - Superior One Bedroom")
		else:
			price1bed_exvat=price1bed.replace(",","")
			nprice1bed=(float(price1bed_exvat))/1.2
			nprice1bed=round(nprice1bed/num_nights)
			print("Superior One Bedroom " + "£ "+ str(nprice1bed))
			sheet.cell(row=thelast+1+i, column=4).value = nprice1bed

		try:
			price2bed = r.html.find("span[id*='mbprice_'][id$='15071']", first=True).text
		except:
			print("No availability - 2 Bedroom Apartment")
		else:
			price2bed_exvat=price2bed.replace(",","")
			nprice2bed=(float(price2bed_exvat))/1.2
			nprice2bed=round(nprice2bed/num_nights)	
			print("2 Bedroom Apartment " + "£ "+ str(nprice2bed))
			sheet.cell(row=thelast+1+i, column=5).value = nprice2bed
		time.sleep(70)
	except Exception as e:
		print(e)
wb.save(filename=destFilename)